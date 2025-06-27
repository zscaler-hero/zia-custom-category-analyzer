#!/usr/bin/env python3

"""
Zscaler ZIA API: Custom URL Category Analyzer (OAuth 2.0 Edition)
Version 1.0 – OAuth 2.0 & OpenAPI Edition

This script authenticates with the ZIA API using OAuth 2.0,
lists your custom URL categories, and analyzes URL coverage for selected categories.

Original script by James Tucker (Zscaler)
OAuth 2.0 version by ZHERO srl
Special thanks to Joost Hage (Zscaler) for valuable contributions and guidance

Outputs:
- % of URLs in known Zscaler categories
- Explicit list of URLs *not* defined/categorized by Zscaler
- Category breakdown
- CSV of all results

HOW TO RUN:
1. Copy .env.example to .env and fill in your OAuth credentials
2. Run:
    python3 zia_custom_category_analyzer_oauth.py
"""

import os
import sys
import time
import requests
import csv
from typing import List, Dict, Tuple, Optional
from dotenv import load_dotenv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load environment variables
load_dotenv()

# Configuration from environment
IDENTITY_BASE_URL = os.getenv("ZSCALER_IDENTITY_BASE_URL")
CLIENT_ID = os.getenv("ZSCALER_CLIENT_ID")
CLIENT_SECRET = os.getenv("ZSCALER_CLIENT_SECRET")


class ZscalerAPIClient:
    """OAuth 2.0 authenticated client for Zscaler ZIA API using Zscaler Identity"""

    def __init__(self, identity_base_url: str, client_id: str, client_secret: str):
        self.identity_base_url = identity_base_url.rstrip("/")
        self.client_id = client_id
        self.client_secret = client_secret
        self.token_url = f"{self.identity_base_url}/oauth2/v1/token"
        self.api_base_url = "https://api.zsapi.net"
        self.session = requests.Session()
        self.access_token = None
        self.token_expiry = 0

        # Rate limiting: 2 seconds between requests (conservative approach)
        self.last_request_time = 0
        self.min_request_interval = 2.0  # seconds

    def _enforce_rate_limit(self):
        """Enforce rate limiting between API calls"""
        current_time = time.time()
        time_since_last_request = current_time - self.last_request_time

        if time_since_last_request < self.min_request_interval:
            sleep_time = self.min_request_interval - time_since_last_request
            time.sleep(sleep_time)

        self.last_request_time = time.time()

    def _get_access_token(self) -> str:
        """Get OAuth 2.0 access token using Zscaler Identity client credentials flow"""

        auth_data = {
            "grant_type": "client_credentials",
            "client_id": self.client_id,
            "client_secret": self.client_secret,
            "audience": "https://api.zscaler.com",
        }

        headers = {"Content-Type": "application/x-www-form-urlencoded"}

        try:
            response = requests.post(
                self.token_url, headers=headers, data=auth_data, timeout=30
            )
            response.raise_for_status()

            data = response.json()
            self.access_token = data.get("access_token")
            # Get actual expiry time from response, default to 1 hour
            expires_in = data.get("expires_in", 3600)
            # Refresh 5 minutes before expiry
            self.token_expiry = time.time() + expires_in - 300

            return self.access_token

        except requests.exceptions.RequestException as e:
            print(f"❌ Authentication failed: {e}")
            if hasattr(e, "response") and e.response is not None:
                print(f"Response: {e.response.text}")
            sys.exit(1)

    def _ensure_authenticated(self):
        """Ensure we have a valid access token"""
        if not self.access_token or time.time() >= self.token_expiry:
            self._get_access_token()
            self.session.headers.update(
                {
                    "Authorization": f"Bearer {self.access_token}",
                    "Content-Type": "application/json",
                }
            )

    def get(self, endpoint: str) -> Dict:
        """Make authenticated GET request with rate limiting and retry on 429"""
        self._ensure_authenticated()
        self._enforce_rate_limit()

        url = f"{self.api_base_url}/zia/api/v1{endpoint}"

        for attempt in range(2):  # Try once, retry once if 429
            try:
                response = self.session.get(url)
                response.raise_for_status()
                return response.json()
            except requests.exceptions.HTTPError as e:
                if e.response.status_code == 429 and attempt == 0:
                    # Handle rate limit - wait and retry once
                    retry_after = 2  # Default to 2 seconds
                    try:
                        # Try to parse Retry-After from response
                        error_data = e.response.json()
                        retry_msg = error_data.get("Retry-After", "2 seconds")
                        retry_after = (
                            int(retry_msg.split()[0]) + 2
                        )  # Add 2 more seconds as requested
                    except:
                        retry_after = 3  # Fallback to 3 seconds

                    print(
                        f"⚠️  Rate limit hit. Waiting {retry_after} seconds before retry..."
                    )
                    time.sleep(retry_after)
                    continue
                else:
                    print(f"❌ API request failed: {e}")
                    if hasattr(e, "response") and e.response is not None:
                        print(f"Response: {e.response.text}")
                    raise
            except requests.exceptions.RequestException as e:
                print(f"❌ API request failed: {e}")
                if hasattr(e, "response") and e.response is not None:
                    print(f"Response: {e.response.text}")
                raise

    def post(self, endpoint: str, data: Dict) -> Dict:
        """Make authenticated POST request with rate limiting and retry on 429"""
        self._ensure_authenticated()
        self._enforce_rate_limit()

        url = f"{self.api_base_url}/zia/api/v1{endpoint}"

        for attempt in range(2):  # Try once, retry once if 429
            try:
                response = self.session.post(url, json=data)
                response.raise_for_status()
                return response.json()
            except requests.exceptions.HTTPError as e:
                if e.response.status_code == 429 and attempt == 0:
                    # Handle rate limit - wait and retry once
                    retry_after = 2  # Default to 2 seconds
                    try:
                        # Try to parse Retry-After from response
                        error_data = e.response.json()
                        retry_msg = error_data.get("Retry-After", "2 seconds")
                        retry_after = (
                            int(retry_msg.split()[0]) + 2
                        )  # Add 2 more seconds as requested
                    except:
                        retry_after = 3  # Fallback to 3 seconds

                    print(
                        f"⚠️  Rate limit hit. Waiting {retry_after} seconds before retry..."
                    )
                    time.sleep(retry_after)
                    continue
                else:
                    print(f"❌ API request failed: {e}")
                    if hasattr(e, "response") and e.response is not None:
                        print(f"Response: {e.response.text}")
                    raise
            except requests.exceptions.RequestException as e:
                print(f"❌ API request failed: {e}")
                if hasattr(e, "response") and e.response is not None:
                    print(f"Response: {e.response.text}")
                raise


def list_custom_categories(client: ZscalerAPIClient) -> List[Dict]:
    """List all custom URL categories"""
    all_categories = client.get("/urlCategories/lite")
    custom_cats = [cat for cat in all_categories if cat.get("customCategory")]

    print(f"\nFound {len(custom_cats)} custom URL categories:")
    print("-" * 60)

    for idx, cat in enumerate(custom_cats, start=1):
        print(f"[{idx}] {cat.get('configuredName')}")
        print(f"     Type: {cat.get('type', 'N/A')}")
        print(f"     Super Category: {cat.get('superCategory', 'N/A')}")
        desc = cat.get("description", "")
        if desc:
            print(f"     Description: {desc}")
        print()

    return custom_cats


def get_category_urls(
    client: ZscalerAPIClient, category_id: str
) -> Tuple[List[str], str]:
    """Get all URLs from a specific category"""
    data = client.get(f"/urlCategories/{category_id}")

    # Combine all URL sources
    urls = []
    urls.extend(data.get("urls", []))
    urls.extend(data.get("dbCategorizedUrls", []))

    # Remove duplicates while preserving order
    seen = set()
    unique_urls = []
    for url in urls:
        if url not in seen:
            seen.add(url)
            unique_urls.append(url)

    return unique_urls, data.get("configuredName", "UNKNOWN")


def lookup_urls(client: ZscalerAPIClient, urls: List[str]) -> List[Dict]:
    """Look up categorization for a list of URLs"""
    all_results = []
    total_chunks = (len(urls) + 99) // 100

    # Zscaler API allows up to 100 URLs per request
    for i in range(0, len(urls), 100):
        chunk_num = (i // 100) + 1
        chunk = urls[i : i + 100]
        print(
            f"  Looking up URLs {i+1}-{min(i+100, len(urls))} of {len(urls)} (chunk {chunk_num}/{total_chunks})..."
        )

        # Rate limiting notice for large batches
        if total_chunks > 5 and chunk_num < total_chunks:
            remaining_time = int((total_chunks - chunk_num) * 2.0)
            print(
                f"    (Rate limited to 2s/request - estimated {remaining_time} seconds remaining)"
            )

        results = client.post("/urlLookup", chunk)
        all_results.extend(results)

    return all_results


def save_to_excel(
    filename: str,
    results_for_csv: List[Dict],
    cat_name: str,
    categorized_count: int,
    uncategorized_urls: List[str],
    category_count: Dict[str, int],
):
    """Save results to Excel file with formatting and credits"""
    wb = Workbook()

    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(
        start_color="1F4788", end_color="1F4788", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Borders
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Uncategorized cell style
    uncategorized_fill = PatternFill(
        start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"
    )

    # Sheet 1: Analysis Results
    ws1 = wb.active
    ws1.title = "Analysis Results"

    # Headers
    headers = ["URL", "Zscaler Categories"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data
    for row_idx, row_data in enumerate(results_for_csv, 2):
        url_cell = ws1.cell(row=row_idx, column=1, value=row_data["url"])
        cat_cell = ws1.cell(row=row_idx, column=2, value=row_data["categories"])

        # Apply borders
        url_cell.border = thin_border
        cat_cell.border = thin_border

        # Highlight uncategorized URLs
        if row_data["categories"] == "<Not categorized>":
            url_cell.fill = uncategorized_fill
            cat_cell.fill = uncategorized_fill

    # Auto-adjust column widths
    for column_cells in ws1.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 100)  # Cap at 100 characters
        ws1.column_dimensions[column].width = adjusted_width

    # Freeze header row
    ws1.freeze_panes = "A2"

    # Sheet 2: Summary
    ws2 = wb.create_sheet(title="Summary")

    # Title
    ws2.merge_cells("A1:D1")
    title_cell = ws2["A1"]
    title_cell.value = f"Analysis Summary: {cat_name}"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center")

    # Date
    ws2.merge_cells("A2:D2")
    date_cell = ws2["A2"]
    date_cell.value = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    date_cell.alignment = Alignment(horizontal="center")

    # Statistics
    ws2["A4"] = "Statistics"
    ws2["A4"].font = Font(bold=True, size=14)

    stats_data = [
        ("Total URLs analyzed:", len(results_for_csv)),
        (
            "URLs categorized by Zscaler:",
            f"{categorized_count} ({round((categorized_count/len(results_for_csv))*100, 2)}%)",
        ),
        (
            "URLs NOT categorized:",
            f"{len(uncategorized_urls)} ({round((len(uncategorized_urls)/len(results_for_csv))*100, 2)}%)",
        ),
    ]

    for idx, (label, value) in enumerate(stats_data, 5):
        ws2[f"A{idx}"] = label
        ws2[f"B{idx}"] = value
        ws2[f"A{idx}"].font = Font(bold=True)

    # Category breakdown
    if category_count:
        ws2["A9"] = "Top Categories"
        ws2["A9"].font = Font(bold=True, size=14)

        # Headers
        ws2["A10"] = "Category"
        ws2["B10"] = "Count"
        ws2["A10"].font = Font(bold=True)
        ws2["B10"].font = Font(bold=True)

        # Data
        sorted_cats = sorted(category_count.items(), key=lambda x: x[1], reverse=True)[
            :15
        ]
        for idx, (cat, count) in enumerate(sorted_cats, 11):
            ws2[f"A{idx}"] = cat
            ws2[f"B{idx}"] = count

    # Credits section
    credits_row = max(26, len(sorted_cats) + 15 if category_count else 15)

    ws2[f"A{credits_row}"] = "Generated by:"
    ws2[f"A{credits_row}"].font = Font(bold=True)

    ws2[f"A{credits_row+1}"] = "Zscaler ZIA Custom URL Category Analyzer"
    ws2[f"A{credits_row+2}"] = "OAuth 2.0 + XLSX Edition by ZHERO srl"

    # Add hyperlink to ZHERO
    ws2[f"B{credits_row+2}"] = "https://zhero.ai"
    ws2[f"B{credits_row+2}"].hyperlink = "https://zhero.ai"
    ws2[f"B{credits_row+2}"].font = Font(color="0000FF", underline="single")

    ws2[f"A{credits_row+3}"] = (
        "Original by James Tucker | Thanks to Joost Hage (Zscaler)"
    )

    # Auto-adjust columns in summary sheet
    for column in ["A", "B", "C", "D"]:
        ws2.column_dimensions[column].width = 40

    # Save the workbook
    wb.save(filename)


def analyze_category(
    client: ZscalerAPIClient, category_id: str, export_format: str = "csv"
):
    """Analyze URL coverage for a specific category"""
    urls, cat_name = get_category_urls(client, category_id)

    print(f"\n{'='*60}")
    print(f"Analyzing category: {cat_name} (ID: {category_id})")
    print(f"{'='*60}")
    print(f"Total URLs in {cat_name}: {len(urls)}")

    if not urls:
        print("No URLs found in this category.")
        return

    print("\nFetching current Zscaler categorization for each URL...")
    lookup_results = lookup_urls(client, urls)

    # Process results
    uncategorized_urls = []
    categorized_count = 0
    category_count = {}
    results_for_csv = []

    for item in lookup_results:
        url = item.get("url")
        categories = item.get("urlClassifications", [])

        if categories:
            categorized_count += 1
            for cat in categories:
                category_count[cat] = category_count.get(cat, 0) + 1
            cat_str = ", ".join(categories)
        else:
            uncategorized_urls.append(url)
            cat_str = "<Not categorized>"

        results_for_csv.append({"url": url, "categories": cat_str})

    # Calculate statistics
    total = len(lookup_results)
    percent_categorized = (
        round((categorized_count / total) * 100, 2) if total > 0 else 0
    )
    percent_uncategorized = (
        round((len(uncategorized_urls) / total) * 100, 2) if total > 0 else 0
    )

    print(f"\n✅ Results:")
    print(
        f"• URLs found in Zscaler-defined categories: {categorized_count} ({percent_categorized}%)"
    )
    print(
        f"• URLs NOT categorized by Zscaler: {len(uncategorized_urls)} ({percent_uncategorized}%)"
    )

    # Show uncategorized URLs
    if uncategorized_urls:
        print(f"\n{'-'*40}")
        print("❗ URLs NOT Defined in Zscaler:")
        for u in uncategorized_urls[:20]:  # Show first 20
            print(f"  - {u}")
        if len(uncategorized_urls) > 20:
            print(f"  ... and {len(uncategorized_urls) - 20} more")
        print(f"{'-'*40}")

    # Show category breakdown
    if category_count:
        print("\nBreakdown by Zscaler category:")
        sorted_cats = sorted(category_count.items(), key=lambda x: x[1], reverse=True)
        for cat, count in sorted_cats[:10]:  # Show top 10
            print(f"  - {cat}: {count} URLs")
        if len(sorted_cats) > 10:
            print(f"  ... and {len(sorted_cats) - 10} more categories")

    # Sample results table
    print("\nSample URL-to-category mapping:")
    print("+---------------------------+-------------------------------+")
    print("| URL                       | Zscaler Category              |")
    print("+---------------------------+-------------------------------+")
    for row in results_for_csv[:7]:
        url_display = row["url"][:26].ljust(26)
        cat_display = row["categories"][:28].ljust(28)
        print(f"| {url_display} | {cat_display} |")
    if len(results_for_csv) > 7:
        print(f"| ... ({len(results_for_csv) - 7} more entries)".ljust(60) + " |")
    print("+---------------------------+-------------------------------+")

    # Save to file based on export format
    if export_format == "xlsx":
        excel_filename = f"{cat_name.lower().replace(' ', '_')}_category_analysis.xlsx"
        save_to_excel(
            excel_filename,
            results_for_csv,
            cat_name,
            categorized_count,
            uncategorized_urls,
            category_count,
        )
        print(f"\n📊 Full URL-to-category mapping saved as: {excel_filename}")
        print("   ✨ Excel file includes formatted sheets with summary and credits")
    else:
        csv_filename = f"{cat_name.lower().replace(' ', '_')}_category_analysis.csv"
        with open(csv_filename, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=["url", "categories"])
            writer.writeheader()
            writer.writerows(results_for_csv)
        print(f"\n📄 Full URL-to-category mapping saved as: {csv_filename}")


def main():
    print("🌐 Zscaler ZIA Custom URL Category Analyzer - OAuth 2.0 Edition")
    print(
        "   Original by James Tucker | OAuth + XLSX version by ZHERO srl | Thanks to Joost Hage"
    )

    # Validate configuration
    if not all([IDENTITY_BASE_URL, CLIENT_ID, CLIENT_SECRET]):
        print("\n❌ Error: Missing OAuth credentials!")
        print("Please copy .env.example to .env and fill in your credentials.")
        sys.exit(1)

    # Initialize API client
    client = ZscalerAPIClient(IDENTITY_BASE_URL, CLIENT_ID, CLIENT_SECRET)

    try:
        # Authenticate
        print(f"\n🔐 Authenticating with Zscaler Identity...")
        client._ensure_authenticated()
        print("✅ Authentication successful!")

        # List custom categories
        print("\n🔍 Fetching custom URL categories...")
        custom_categories = list_custom_categories(client)

        if not custom_categories:
            print("No custom URL categories found.")
            return

        # Get user selection
        print("\n📋 Select categories to analyze:")
        selection = input(
            "Enter category numbers (e.g., 1,3,5) or 'all' for all categories: "
        ).strip()

        selected_ids = []
        if selection.lower() == "all":
            selected_ids = [cat.get("id") for cat in custom_categories]
        else:
            for s in selection.split(","):
                try:
                    idx = int(s.strip()) - 1
                    if 0 <= idx < len(custom_categories):
                        selected_ids.append(custom_categories[idx].get("id"))
                    else:
                        print(f"⚠️  Skipping invalid selection: {s}")
                except ValueError:
                    print(f"⚠️  Skipping invalid selection: {s}")

        if not selected_ids:
            print("No valid categories selected.")
            return

        # Get export format preference
        print("\n📄 Select export format:")
        print("1. CSV (Comma-separated values)")
        print("2. XLSX (Excel with formatting)")
        export_choice = input("Enter your choice (1 or 2) [default: 2]: ").strip()

        export_format = "xlsx"  # Default to Excel
        if export_choice == "1":
            export_format = "csv"
        elif export_choice != "2" and export_choice != "":
            print("Invalid choice. Using Excel format.")

        # Analyze selected categories
        print(f"\n🔎 Analyzing {len(selected_ids)} categories...")
        for cat_id in selected_ids:
            try:
                analyze_category(client, cat_id, export_format)
            except Exception as e:
                print(f"\n❌ Error analyzing category {cat_id}: {e}")
                continue

        print("\n✅ Analysis complete!")

    except KeyboardInterrupt:
        print("\n\n⚠️  Analysis interrupted by user.")
    except Exception as e:
        print(f"\n❌ Unexpected error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
